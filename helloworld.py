from openpyxl import Workbook

wb = Workbook()
ws = wb.active

ws['A1'] = 'hello'
ws['B1'] = 'world!'

wb.save(filename='hello_world.xlsx')